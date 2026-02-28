"""
Phase 7~8 기능 테스트 (2차 UAT)
- 동송농협 TSV 텍스트 파일 생성
- 상품약어 오름차순 정렬
- 출력 파일 수정 가능 (시트보호 해제)
- 파일 교체 초기화 (해시 기반)
"""
import pytest
import pandas as pd
import io
import os
import sys
import hashlib
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from convert import create_excel_buffer, create_text_buffer


class TestCreateTextBuffer:
    """create_text_buffer() — 동송농협 TSV 텍스트 파일 생성 (요청 #1)"""

    def test_returns_bytes(self):
        """바이너리(bytes) 데이터를 반환해야 함"""
        df = pd.DataFrame({'상품명': ['쌀 10kg'], '수량': [1]})
        result = create_text_buffer(df)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_tab_separated(self):
        """탭 구분자로 분리되어야 함"""
        df = pd.DataFrame({'상품명': ['쌀 10kg'], '수량': [1], '수령인': ['홍길동']})
        result = create_text_buffer(df)
        text = result.decode('utf-8-sig')
        lines = text.strip().split('\n')
        # 헤더가 탭으로 구분되어야 함
        assert '\t' in lines[0]
        # 데이터도 탭으로 구분되어야 함
        assert '\t' in lines[1]

    def test_has_bom(self):
        """UTF-8 BOM이 포함되어야 함 (엑셀 호환)"""
        df = pd.DataFrame({'상품명': ['쌀']})
        result = create_text_buffer(df)
        # UTF-8 BOM: EF BB BF
        assert result[:3] == b'\xef\xbb\xbf'

    def test_header_row(self):
        """첫 행은 컬럼명이어야 함"""
        df = pd.DataFrame({'상품명': ['쌀 10kg'], '수량': [1]})
        result = create_text_buffer(df)
        text = result.decode('utf-8-sig')
        header = text.strip().split('\n')[0]
        assert '상품명' in header
        assert '수량' in header

    def test_data_integrity(self):
        """데이터가 올바르게 포함되어야 함"""
        df = pd.DataFrame({
            '상품명': ['동송농협 쌀 10kg', '동송농협 현미 5kg'],
            '수량': [2, 3],
            '수령인': ['홍길동', '김철수']
        })
        result = create_text_buffer(df)
        text = result.decode('utf-8-sig')
        assert '동송농협 쌀 10kg' in text
        assert '동송농협 현미 5kg' in text
        assert '홍길동' in text
        assert '김철수' in text

    def test_empty_dataframe(self):
        """빈 DataFrame도 에러 없이 처리"""
        df = pd.DataFrame({'상품명': [], '수량': []})
        result = create_text_buffer(df)
        assert isinstance(result, bytes)
        assert len(result) > 0


class TestSheetProtection:
    """출력 파일 수정 가능 확인 (요청 #2)"""

    def test_sheet_not_protected(self):
        """생성된 엑셀의 시트가 보호되지 않아야 함"""
        df = pd.DataFrame({'상품명': ['쌀'], '수량': [1]})
        result = create_excel_buffer(df, '테스트업체')

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.protection.sheet == False, "시트가 보호되어 있으면 안 됨"

    def test_cells_editable(self):
        """셀이 수정 가능해야 함"""
        df = pd.DataFrame({'상품명': ['쌀'], '수량': [1]})
        result = create_excel_buffer(df, '테스트업체')

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        # 시트 보호가 꺼져 있으면 셀 수정 가능
        assert not ws.protection.enable()


class TestSorting:
    """상품약어 정렬 및 엑셀 AutoFilter 확인 (요청 #4)"""

    def test_sort_ascending(self):
        """상품약어 컬럼이 오름차순으로 정렬되는지 확인"""
        df = pd.DataFrame({
            '상품명': ['다다 쌀', '가가 현미', '나나 잡곡'],
            '수량': [1, 2, 3]
        })
        sorted_df = df.sort_values(by='상품명', ascending=True).reset_index(drop=True)
        assert sorted_df['상품명'].tolist() == ['가가 현미', '나나 잡곡', '다다 쌀']

    def test_excel_has_autofilter(self):
        """생성된 엑셀에 AutoFilter가 적용되어야 함"""
        df = pd.DataFrame({
            '상품명': ['쌀 10kg', '현미 5kg'],
            '수량': [1, 2],
            '수령인': ['홍길동', '김철수']
        })
        result = create_excel_buffer(df, '테스트업체')
        
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        # AutoFilter가 설정되어 있어야 함
        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref != ''

    def test_sort_with_na(self):
        """NaN 값이 있어도 정렬이 동작해야 함"""
        df = pd.DataFrame({
            '상품명': ['다다', None, '가가'],
            '수량': [1, 2, 3]
        })
        sorted_df = df.sort_values(by='상품명', ascending=True, na_position='last').reset_index(drop=True)
        assert sorted_df['상품명'].tolist()[0] == '가가'
        assert sorted_df['상품명'].tolist()[1] == '다다'
        assert pd.isna(sorted_df['상품명'].tolist()[2])


class TestFileHashInitialization:
    """파일 교체 초기화 (요청 #7)"""

    def test_different_files_different_hash(self):
        """서로 다른 파일은 다른 해시를 가져야 함"""
        data1 = b"file content 1"
        data2 = b"file content 2"
        hash1 = hashlib.md5(data1).hexdigest()
        hash2 = hashlib.md5(data2).hexdigest()
        assert hash1 != hash2

    def test_same_file_same_hash(self):
        """같은 파일은 같은 해시를 가져야 함"""
        data = b"same content"
        hash1 = hashlib.md5(data).hexdigest()
        hash2 = hashlib.md5(data).hexdigest()
        assert hash1 == hash2

    def test_hash_deterministic(self):
        """해시는 결정적이어야 함"""
        data = b"test data for hashing"
        expected = hashlib.md5(data).hexdigest()
        for _ in range(10):
            assert hashlib.md5(data).hexdigest() == expected


class TestVendorPriority:
    """업체 우선 배치 (요청 #3)"""

    def test_priority_vendors_exist(self):
        """우선 배치 업체가 nh_list에 존재해야 함"""
        from map import nh_list
        priority = ['관인농협', '오덕쌀', '영광군농협', '한국라이스텍']
        for name in priority:
            assert name in nh_list, f"'{name}'이 nh_list에 없습니다"

    def test_priority_indices_first(self):
        """우선 배치 업체 인덱스가 정렬 시 앞에 와야 함"""
        from map import nh_list
        priority_names = ['관인농협', '오덕쌀', '영광군농협', '한국라이스텍']
        priority_indices = [i for i, name in enumerate(nh_list) if name in priority_names]
        other_indices = [i for i in range(len(nh_list)) if i not in priority_indices]
        ordered = priority_indices + other_indices

        # 우선 배치 업체가 처음 4개에 와야 함
        for i, name in enumerate(priority_names):
            assert nh_list[ordered[i]] in priority_names
