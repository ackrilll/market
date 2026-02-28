import json
import pandas as pd
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONFIG_PATH = os.path.join(BASE_DIR, "data", "vendor_config.json")
OUTPUT_PATH = os.path.join(BASE_DIR, "user_test", "업체별_분류_키워드_목록.xlsx")

with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    config = json.load(f)

rows = []
for v in config["vendors"]:
    rename_count = len(v.get("rename_map", {}))
    target_count = len(v.get("target_columns", []))
    const_count = len(v.get("constant_values", {}))
    copy_count = len(v.get("copy_map", {}))
    has_split = "O" if v.get("split_config") else ""
    rename_type = "고유" if rename_count > 0 else "기본(default)"

    rows.append({
        "idx": v["id"],
        "업체명": v["name"],
        "분류 키워드": v["name"],
        "매칭 방식": "str.contains (부분 문자열)",
        "rename_map": rename_type,
        "target_columns 수": target_count,
        "constant_values 수": const_count,
        "copy_map 수": copy_count,
        "파일 분할": has_split,
    })

df = pd.DataFrame(rows)

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="분류 키워드")
    ws = writer.sheets["분류 키워드"]

    # 스타일링
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="맑은 고딕", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 헤더 스타일
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 데이터 스타일
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # 컬럼 너비
    col_widths = [6, 18, 18, 28, 16, 18, 18, 14, 12]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = w

    # 필터
    ws.auto_filter.ref = ws.dimensions

print(f"OK: {OUTPUT_PATH}")
